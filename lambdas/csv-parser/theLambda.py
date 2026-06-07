import boto3
import io
import logging
import json
import email
import os
import zipfile
import xml.etree.ElementTree as ET
import csv
import openpyxl
import random
import time
from openpyxl.utils import column_index_from_string
from datetime import datetime

s3 = boto3.client("s3")
sns_client = boto3.client("sns")
ses_client = boto3.client("ses")

logger = logging.getLogger()
logger.setLevel(logging.INFO)

CSV_BUCKET   = os.environ.get("CSV_BUCKET")
SENDER_EMAIL = os.environ.get("SENDER_EMAIL")
NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _decode(val):
    return (val or "").replace("*", " ")


FILE_PREFIX_UNFILLED        = _decode(os.environ.get("FILE_PREFIX_UNFILLED"))
FILE_PREFIX_CONTRACTOR_OPEN = _decode(os.environ.get("FILE_PREFIX_CONTRACTOR_OPEN"))
FILE_PREFIX_CANDIDATES      = _decode(os.environ.get("FILE_PREFIX_CANDIDATES"))
# Base contractor prefix — same workbook, two sheets (Open/Closed)
FILE_PREFIX_CONTRACTOR_BASE = FILE_PREFIX_CONTRACTOR_OPEN.rsplit("-Open", 1)[0]


def process_file(filename, file_content):
    try:
        if not filename.lower().endswith(".xlsx"):
            logger.info("Skipping unsupported file in process_file: %s", filename)
            return []

        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
        base_name = filename.rsplit(".", 1)[0]
        clean_base = base_name.split(" 2")[0].split("-2")[0].rstrip("- ")

        results = []

        if filename.startswith("NEW-IT Contractor-VG-Vendor Req Report"):
            for sheet in ("Open", "Closed"):
                logger.info("Processing sheet '%s' (NEW-IT) from: %s", sheet, filename)

                # Extract the sheet into a standalone XLSX
                sheet_bytes = extract_sheet_to_xlsx_bytes(file_content, sheet)
                if not sheet_bytes:
                    logger.warning("No rows or sheet '%s' not found in %s", sheet, filename)
                    continue

                xlsx_key = f"{clean_base}-{sheet}-{timestamp}.xlsx"
                s3.put_object(
                    Bucket=CSV_BUCKET,
                    Key=xlsx_key,
                    Body=sheet_bytes,
                    ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                logger.info("Uploaded XLSX (NEW-IT): %s/%s", CSV_BUCKET, xlsx_key)
                results.append(xlsx_key)

            return results

        # All other reports: copy as-is with a timestamp so rollcall-lambda can find the newest by prefix
        copy_key = f"{clean_base}-copy-{timestamp}.xlsx"
        s3.put_object(
            Bucket=CSV_BUCKET,
            Key=copy_key,
            Body=file_content,
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        logger.info("Uploaded copy: %s/%s", CSV_BUCKET, copy_key)
        results.append(copy_key)
        return results

    except Exception as e:
        logger.error("Error in process_file %s: %s", filename, e)
        return []
    
def extract_sheet_to_xlsx_bytes(xlsx_bytes, sheet_name):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
        if sheet_name not in wb.sheetnames:
            return None

        src = wb[sheet_name]
        new_wb = openpyxl.Workbook()
        dst = new_wb.active
        dst.title = src.title

        for row in src.iter_rows():
            for cell in row:
                # col_idx may not exist on merged/special cells; fall back to conversion
                col_idx = getattr(cell, 'col_idx', None)
                if col_idx is None:
                    try:
                        col_idx = column_index_from_string(cell.column)
                    except Exception:
                        continue

                dcell = dst.cell(row=cell.row, column=col_idx, value=cell.value)
                try:
                    dcell.number_format = cell.number_format
                    if getattr(cell, 'has_style', False):
                        dcell.font = cell.font
                        dcell.fill = cell.fill
                        dcell.border = cell.border
                        dcell.alignment = cell.alignment
                        dcell.protection = cell.protection
                except Exception:
                    pass

        try:
            for col_letter, dim in src.column_dimensions.items():
                if dim.width is not None:
                    new_wb.active.column_dimensions[col_letter].width = dim.width
        except Exception:
            pass

        try:
            for idx, dim in src.row_dimensions.items():
                if getattr(dim, 'height', None) is not None:
                    new_wb.active.row_dimensions[idx].height = dim.height
        except Exception:
            pass

        try:
            for merged in src.merged_cells.ranges:
                dst.merge_cells(str(merged))
        except Exception:
            pass

        try:
            if src.freeze_panes:
                dst.freeze_panes = src.freeze_panes
        except Exception:
            pass

        buf = io.BytesIO()
        new_wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    except Exception as e:
        logger.error("Error extracting sheet %s: %s", sheet_name, e)
        return None




def col_letter_to_index(col):
    index = 0
    for char in col:
        index = index * 26 + (ord(char.upper()) - ord('A') + 1)
    return index - 1



def parse_csv_to_rows(csv_bytes):
    try:
        text = csv_bytes.decode("utf-8")
    except UnicodeDecodeError:
        text = csv_bytes.decode("latin1")

    reader = csv.reader(io.StringIO(text))
    return list(reader)



def publish_to_sns(ret_addr: str, orig_message_id: str = "", orig_subject: str = ""):
    topic_arn = os.environ.get("SNS_TOPIC_ARN")
    message = {
        "status": "processed",
        "bucket": CSV_BUCKET,
        "retAddr": ret_addr,
        "origMessageId": orig_message_id,
        "origSubject": orig_subject,
    }
    response = sns_client.publish(
        TopicArn=topic_arn,
        Message=json.dumps(message),
        Subject="New XLSX File Processed"
    )
    logger.info("Published SNS message: %s", response["MessageId"])


def _send_failure_email(to_email, subject_ref="", error_msg=""):
    if not to_email or not SENDER_EMAIL:
        logger.warning("Cannot send failure notification — sender or recipient email not configured")
        return
    try:
        ref = f' "{subject_ref}"' if subject_ref else ""
        body = (
            f"Your submission{ref} could not be processed. An error occurred while "
            "reading the attached files and no output report was generated.\n\n"
        )
        if error_msg:
            body += f"Error: {error_msg}\n\n"
        body += (
            "Please check that all required report files are attached and resubmit. "
            "If the problem continues, contact your administrator."
        )
        ses_client.send_email(
            Source=SENDER_EMAIL,
            Destination={"ToAddresses": [to_email]},
            Message={
                "Subject": {"Data": "Pipeline Error — Unable to Process Submission"},
                "Body":    {"Text": {"Data": body}},
            },
        )
        logger.info("Failure notification sent to %s", to_email)
    except Exception as e:
        logger.error("Could not send failure notification to %s: %s", to_email, e)


def _get_object_with_retry(bucket, key, max_retries=3):
    for attempt in range(max_retries):
        try:
            return s3.get_object(Bucket=bucket, Key=key)
        except s3.exceptions.NoSuchKey:
            raise
        except Exception as e:
            if attempt < max_retries - 1:
                wait = (2 ** attempt) + random.uniform(0, 1)
                logger.warning("S3 get_object retry %d/%d in %.1fs: %s", attempt + 1, max_retries - 1, wait, e)
                time.sleep(wait)
            else:
                raise


def wipe_buckets():
    s3_resource = boto3.resource("s3")
    bucket = s3_resource.Bucket(CSV_BUCKET)
    response = bucket.objects.all().delete()
    count = len(response[0].get("Deleted", [])) if response else 0
    logger.info("Wiped %d object(s) from %s", count, CSV_BUCKET)



def lambda_handler(event, context):
    start = time.time()
    logger.info(f"Event received: {json.dumps(event, indent=2)}")

    sender_email    = None
    orig_message_id = ""
    orig_subject    = ""
    try:
        wipe_buckets()

        first_msg       = json.loads(json.loads(event["Records"][0]["body"])["Message"])
        sender_email    = first_msg["mail"]["source"]
        common_headers  = first_msg.get("mail", {}).get("commonHeaders", {})
        orig_message_id = common_headers.get("messageId", "")
        orig_subject    = common_headers.get("subject", "")
        logger.info("Email source: %s | orig messageId: %s | subject: %s", sender_email, orig_message_id, orig_subject)

        BLOCKED_SENDERS = {"no-reply-aws@amazon.com"}
        if sender_email in BLOCKED_SENDERS:
            logger.info(f"Ignoring email from blocked sender: {sender_email}")
            return {"status": "ignored", "reason": "blocked sender"}

        total_files = 0
        processed_files = []
        seen_filenames = []

        for record in event.get("Records", []):
            sns_message = record["body"]
            message = json.loads(sns_message)["Message"]
            mail_obj = json.loads(message)

            s3_info = mail_obj["receipt"]["action"]
            bucket_name = s3_info["bucketName"]
            object_key = s3_info["objectKey"]

            logger.info("Processing email: %s/%s", bucket_name, object_key)

            email_obj = _get_object_with_retry(bucket_name, object_key)
            msg = email.message_from_binary_file(email_obj["Body"])

            for part in msg.walk():
                if "attachment" not in part.get("Content-Disposition", ""):
                    continue
                if part.get_content_maintype() == "multipart":
                    continue

                filename = part.get_filename()
                if not filename:
                    logger.info("Skipping attachment part with no filename (Content-Type: %s)", part.get_content_type())
                    continue

                file_content = part.get_payload(decode=True)
                total_files += 1

                results = process_file(filename, file_content)
                if results:
                    seen_filenames.append(filename)
                processed_files.extend(results)

                del file_content

        logger.info("Processed %d attachment(s) → %d output file(s): %s", total_files, len(processed_files), processed_files)

        missing = []
        if not any(f.startswith(FILE_PREFIX_UNFILLED) for f in seen_filenames):
            missing.append(f"unfilled report (expected prefix: '{FILE_PREFIX_UNFILLED}')")
        if not any(f.startswith(FILE_PREFIX_CONTRACTOR_BASE) for f in seen_filenames):
            missing.append(f"contractor report (expected prefix: '{FILE_PREFIX_CONTRACTOR_BASE}')")
        if not any(f.startswith(FILE_PREFIX_CANDIDATES) for f in seen_filenames):
            missing.append(f"candidates report (expected prefix: '{FILE_PREFIX_CANDIDATES}')")
        if missing:
            raise ValueError(f"Email is missing required attachments: {'; '.join(missing)}")

        publish_to_sns(sender_email, orig_message_id, orig_subject)

        elapsed = time.time() - start
        logger.info("Handler complete in %.2fs", elapsed)
        return {
            "status": "success",
            "processed_files": processed_files,
            "total_files": total_files
        }

    except Exception as e:
        logger.error("Unhandled exception: %s", e, exc_info=True)
        _send_failure_email(sender_email, orig_subject, error_msg=str(e))
        raise